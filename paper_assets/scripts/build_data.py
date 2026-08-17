#!/usr/bin/env python3
"""Aggregate the raw experiment tree of the code repository into the small, tidy
CSVs that this LaTeX project ships and that every figure/table script reads.

These aggregates live here, in the code repository, next to the raw results they
come from -- the paper repository carries only what a compile needs, and a
compile never reads a CSV. They are provenance, not build input: this script
documents where every number in the manuscript comes from and regenerates them
byte-for-byte.

Source repository (override with $HCMAGRL_CODE):
    /home/user/code-Junxin-Huang-HCMADRL

Outputs (all under figures/data/):
    convergence_<group>.csv   normalized learning curves, mean +/- std over instances
    eval_cells.csv            per (method, instance, weight) mean/std over 20 seeds
    cell_scores.csv           min-max normalized scalarized score + rank per cell
    pareto_points.csv         every evaluation, with the non-dominated ones flagged
    stats_summary.csv         paired Wilcoxon, Cohen's d, average ranks, Nemenyi CD
    indicators.csv            IGD / Spread / HV per instance (from HCMADRL.xlsx)
    indicator_winrate.csv     per-indicator win rates (from HCMADRL.xlsx)
    benchmark_rmssp.csv       RMSSP makespan vs metaheuristic baselines
    runtime.csv               mean seconds per training episode by method and size
    instances.csv             the 27 generated instances and their dimensions
"""
from __future__ import annotations

import ast
import os
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats

CODE = Path(os.environ.get("HCMAGRL_CODE", "/home/user/code-Junxin-Huang-HCMADRL"))
OUT = Path(__file__).resolve().parent.parent / "figures" / "data"

# Directory name in result/ -> name used throughout the paper.
METHODS = {
    "HCMADRL": "HCMAGRL",
    "No_Attn": "w/o EdgeAttn",
    "Flat_DRL": "w/o Hierarchy",
    "HomoGNN": "w/o Heterogeneity",
    "MLP_Encoder": "w/o Graph",
    "DDQN": "DDQN",
    "EDQN": "EDQN",
    "SAC": "SAC",
    "TD3": "D-DRL",
}
OURS = "HCMAGRL"
ABLATION = [OURS, "w/o EdgeAttn", "w/o Hierarchy", "w/o Heterogeneity", "w/o Graph"]
RL_BASELINES = [OURS, "DDQN", "EDQN", "SAC", "D-DRL"]

# config.py:26-32 -- fixes the C1..C27 labelling used by HCMADRL.xlsx.
INSTANCE_ORDER = [
    "M4_A3_R4_J2", "M4_A3_R6_J3", "M4_A3_R8_J4", "M4_A4_R4_J2", "M4_A4_R6_J3",
    "M4_A4_R8_J4", "M4_A5_R4_J2", "M4_A5_R6_J3", "M4_A5_R8_J4", "M8_A6_R4_J2",
    "M8_A6_R6_J3", "M8_A6_R8_J4", "M8_A8_R4_J2", "M8_A8_R6_J3", "M8_A8_R8_J4",
    "M8_A10_R4_J2", "M8_A10_R6_J3", "M8_A10_R8_J4", "M12_A9_R4_J2", "M12_A9_R6_J3",
    "M12_A9_R8_J4", "M12_A12_R4_J2", "M12_A12_R6_J3", "M12_A12_R8_J4",
    "M12_A15_R4_J2", "M12_A15_R6_J3", "M12_A15_R8_J4",
]
CID = {name: f"C{i + 1}" for i, name in enumerate(INSTANCE_ORDER)}

# Nemenyi two-tailed critical values q_alpha at alpha=0.05, indexed by k.
NEMENYI_Q05 = {
    2: 1.960, 3: 2.343, 4: 2.569, 5: 2.728, 6: 2.850, 7: 2.949,
    8: 3.031, 9: 3.102, 10: 3.164,
}


def parse_instance(name: str) -> dict:
    m = re.fullmatch(r"M(\d+)_A(\d+)_R(\d+)_J(\d+)", name)
    if not m:
        raise ValueError(f"unparseable instance name: {name}")
    M, A, R, J = (int(x) for x in m.groups())
    return {"instance_id": name, "M": M, "A": A, "R": R, "J": J, "op_types": R * J}


# --------------------------------------------------------------------------
# 1. Convergence curves
# --------------------------------------------------------------------------
def build_convergence() -> None:
    """Each run is normalized by its own first-episode value before averaging.

    Raw makespans span 960..3000+ across the 27 instances, so an unnormalized
    mean would simply track the largest instances. Normalizing per run makes the
    curves comparable and the shaded band meaningful (spread across instances).
    """
    rows = []
    for src, label in METHODS.items():
        for weight, wdir in (("balanced", "wt0.5_wc0.5"), ("makespan", "wt1_wc0")):
            curves = {"makespan_mean": [], "cost_mean": [], "steps_mean": []}
            for inst in INSTANCE_ORDER:
                p = CODE / "result" / src / f"{src}_{wdir}_{inst}" / "log.csv"
                if not p.exists():
                    continue
                d = pd.read_csv(p)
                for col in curves:
                    v = d[col].to_numpy(dtype=float)
                    base = v[0] if v[0] not in (0.0, np.nan) else np.nan
                    curves[col].append(v / base if base and np.isfinite(base) else np.full_like(v, np.nan))
            if not curves["makespan_mean"]:
                continue
            n_inst = len(curves["makespan_mean"])
            for col in curves:
                arr = np.vstack(curves[col])
                with np.errstate(invalid="ignore"):
                    mean = np.nanmean(arr, axis=0)
                    std = np.nanstd(arr, axis=0)
                for ep, (mu, sd) in enumerate(zip(mean, std), start=1):
                    rows.append({
                        "method": label, "weight": weight, "metric": col,
                        "episode": ep, "mean": mu, "std": sd, "n_instances": n_inst,
                    })
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "convergence.csv", index=False)
    print(f"  convergence.csv          {len(df):>6} rows")


# --------------------------------------------------------------------------
# 2. Evaluation cells and scalarized scores
# --------------------------------------------------------------------------
def build_eval() -> pd.DataFrame:
    d = pd.read_csv(CODE / "result" / "eval" / "eval_detail.csv")
    d["method"] = d["method"].map(METHODS)
    cells = (d.groupby(["method", "instance_id", "alpha_t", "alpha_c"])
               .agg(makespan_mean=("makespan", "mean"), makespan_std=("makespan", "std"),
                    cost_mean=("cost", "mean"), cost_std=("cost", "std"),
                    n_seeds=("seed", "nunique"))
               .reset_index())
    cells["cid"] = cells["instance_id"].map(CID)
    cells.to_csv(OUT / "eval_cells.csv", index=False)
    print(f"  eval_cells.csv           {len(cells):>6} rows")

    # Within each (instance, weight) cell the two objectives are min-max scaled
    # across the 9 methods, then combined with that cell's own preference
    # weights. This makes the 135 cells commensurable for ranking.
    scored = cells.copy()
    cell = scored.groupby(["instance_id", "alpha_t"])
    for obj in ("makespan", "cost"):
        lo = cell[f"{obj}_mean"].transform("min")
        hi = cell[f"{obj}_mean"].transform("max")
        rng = hi - lo
        scored[f"{obj}_norm"] = np.where(rng == 0, 0.0,
                                         (scored[f"{obj}_mean"] - lo) / rng.replace(0, np.nan))
        scored[f"{obj}_norm"] = scored[f"{obj}_norm"].fillna(0.0)
    scored["score"] = (scored["alpha_t"] * scored["makespan_norm"]
                       + scored["alpha_c"] * scored["cost_norm"])
    scored["rank"] = scored.groupby(["instance_id", "alpha_t"])["score"].rank(method="average")
    scored.to_csv(OUT / "cell_scores.csv", index=False)
    print(f"  cell_scores.csv          {len(scored):>6} rows")
    return scored


def nondominated(points: np.ndarray) -> np.ndarray:
    """Mask of the non-dominated rows of an (n, 2) array, both objectives minimized.

    A point is dominated when another point is no worse on both objectives and
    strictly better on at least one. n is 100 here, so the quadratic sweep costs
    nothing and stays readable.
    """
    keep = np.ones(len(points), dtype=bool)
    for i, p in enumerate(points):
        if not keep[i]:
            continue
        beaten = np.all(points <= p, axis=1) & np.any(points < p, axis=1)
        if beaten.any():
            keep[i] = False
    return keep


def build_pareto() -> None:
    """Every evaluation behind the weight sweep, with its non-dominated ones flagged.

    The indicator workbook scores the five weight-swept operating points; the
    front figures need the evaluations underneath them, because a front drawn
    from five means collapses to one or two points on most instances. Each
    (method, instance) therefore contributes all 5 x 20 = 100 evaluations and
    the front is the non-dominated subset of those.
    """
    d = pd.read_csv(CODE / "result" / "eval" / "eval_detail.csv")
    d["method"] = d["method"].map(METHODS)
    d["cid"] = d["instance_id"].map(CID)
    d = d[["method", "cid", "instance_id", "alpha_t", "seed", "makespan", "cost"]]

    flags = np.zeros(len(d), dtype=bool)
    for _, idx in d.groupby(["method", "instance_id"], sort=False).indices.items():
        flags[idx] = nondominated(d.iloc[idx][["makespan", "cost"]].to_numpy())
    d["nondominated"] = flags

    d.to_csv(OUT / "pareto_points.csv", index=False)
    sizes = d[d.nondominated].groupby(["method", "instance_id"]).size()
    print(f"  pareto_points.csv        {len(d):>6} rows "
          f"(front size {sizes.min()}-{sizes.max()}, median {sizes.median():.0f})")


# --------------------------------------------------------------------------
# 3. Statistics: paired Wilcoxon, Cohen's d, Friedman + Nemenyi
# --------------------------------------------------------------------------
def build_stats(scored: pd.DataFrame) -> None:
    key = ["instance_id", "alpha_t"]
    wide = scored.pivot_table(index=key, columns="method", values="score")
    m_wide = scored.pivot_table(index=key, columns="method", values="makespan_mean")
    c_wide = scored.pivot_table(index=key, columns="method", values="cost_mean")

    rows = []
    for metric, table, better in (("score", wide, "lower"),
                                  ("makespan", m_wide, "lower"),
                                  ("cost", c_wide, "lower")):
        a = table[OURS]
        for other in table.columns:
            if other == OURS:
                continue
            b = table[other]
            diff = (a - b).dropna()
            if diff.empty or np.allclose(diff, 0):
                continue
            w_stat, p = stats.wilcoxon(a, b)
            # Paired Cohen's d (mean difference over sd of the differences).
            d_val = diff.mean() / diff.std(ddof=1) if diff.std(ddof=1) else np.nan
            rel = 100.0 * (a.mean() - b.mean()) / b.mean() if b.mean() else np.nan
            rows.append({
                "metric": metric, "method_a": OURS, "method_b": other,
                "mean_a": a.mean(), "mean_b": b.mean(), "rel_gap_pct": rel,
                "statistic": w_stat, "p_value": p, "cohens_d": d_val,
                "n_pairs": len(diff), "better": better,
            })
    pairwise = pd.DataFrame(rows)
    pairwise.to_csv(OUT / "stats_pairwise.csv", index=False)
    print(f"  stats_pairwise.csv       {len(pairwise):>6} rows")

    def friedman_nemenyi(sub: pd.DataFrame, family: str) -> pd.DataFrame:
        """Rank within the given family, then Friedman with the Nemenyi CD.

        The comparison families are adjudicated separately. The ablation study
        and the baseline comparison answer different questions and are not
        compared with one another, so correcting across their union would only
        inflate the critical difference without controlling any error rate we
        actually care about.
        """
        s = sub.copy()
        s["r"] = s.groupby(key)["score"].rank(method="average")
        w = s.pivot_table(index=key, columns="method", values="r").dropna()
        k, N = w.shape[1], len(w)
        chi2, p = stats.friedmanchisquare(*[w[c].to_numpy() for c in w.columns])
        cd = NEMENYI_Q05[k] * np.sqrt(k * (k + 1) / (6.0 * N))
        out = (w.mean().rename("avg_rank").reset_index()
                .sort_values("avg_rank").reset_index(drop=True))
        out.insert(0, "family", family)
        out["k_methods"], out["n_blocks"] = k, N
        out["friedman_chi2"], out["friedman_p"], out["nemenyi_cd"] = chi2, p, cd
        best = out["avg_rank"].iloc[0]
        out["gap_to_best"] = out["avg_rank"] - best
        out["separated"] = out["gap_to_best"] > cd
        print(f"  [{family}] k={k} N={N} chi2={chi2:.1f} p={p:.3g} CD={cd:.3f}")
        return out

    families = {
        "ablation": [OURS] + ABLATION[1:],
        "baseline": [OURS] + RL_BASELINES[1:],
    }
    parts = [friedman_nemenyi(scored[scored.method.isin(ms)], fam)
             for fam, ms in families.items()]
    parts.append(friedman_nemenyi(scored, "all"))
    summary = pd.concat(parts, ignore_index=True)
    summary.to_csv(OUT / "stats_summary.csv", index=False)
    print(f"  stats_summary.csv        {len(summary):>6} rows")


# --------------------------------------------------------------------------
# 4. Multi-objective indicators from HCMADRL.xlsx
# --------------------------------------------------------------------------
def build_indicators() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(CODE / "result" / "HCMADRL.xlsx", data_only=True)
    blocks = {"IGD": 2, "Spread": 8, "HV": 14}  # 1-based first data column
    rows, wins = [], []
    for sheet, group in (("RL", "rl"), ("Ablation", "ablation")):
        ws = wb[sheet]
        for ind, c0 in blocks.items():
            names = [ws.cell(2, c).value for c in range(c0, c0 + 5)]
            for r in range(3, 30):  # C1..C27
                cid = ws.cell(r, 1).value
                for off, nm in enumerate(names):
                    rows.append({
                        "group": group, "indicator": ind, "cid": cid,
                        "method": METHODS.get(nm, nm), "value": ws.cell(r, c0 + off).value,
                    })
        wnames = [ws.cell(33, c).value for c in range(2, 7)]
        for r, ind in ((34, "IGD"), (35, "Spread"), (36, "HV")):
            for off, nm in enumerate(wnames):
                raw = ws.cell(r, 2 + off).value
                wins.append({
                    "group": group, "indicator": ind,
                    "method": METHODS.get(nm, nm),
                    "win_rate": float(str(raw).rstrip("%")),
                })
    pd.DataFrame(rows).to_csv(OUT / "indicators.csv", index=False)
    pd.DataFrame(wins).to_csv(OUT / "indicator_winrate.csv", index=False)
    print(f"  indicators.csv           {len(rows):>6} rows")
    print(f"  indicator_winrate.csv    {len(wins):>6} rows")


# --------------------------------------------------------------------------
# 5. Public RMSSP benchmark vs metaheuristics
# --------------------------------------------------------------------------
def build_benchmark() -> None:
    import openpyxl
    ws = openpyxl.load_workbook(CODE / "result" / "HCMADRL.xlsx", data_only=True)["VS DABC"]
    sizes = [str(ws.cell(2, c).value).replace("×", "x") for c in range(2, 14)]
    rows = []
    for criterion, r0 in (("evaluations", 3), ("runtime", 12)):
        for r in range(r0, r0 + 5):
            algo = str(ws.cell(r, 1).value).split("(")[0].strip()
            algo = OURS if algo == "HGNN" else algo
            for off, size in enumerate(sizes):
                rows.append({
                    "criterion": criterion, "algorithm": algo, "size": size,
                    "makespan": float(ws.cell(r, 2 + off).value),
                })
    # case0..case4: the five independent runs behind our own averaged row.
    for r in range(19, 24):
        case = ws.cell(r, 1).value
        for off, size in enumerate(sizes):
            rows.append({
                "criterion": "cases", "algorithm": f"{OURS}-{case}", "size": size,
                "makespan": float(ws.cell(r, 2 + off).value),
            })
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "benchmark_rmssp.csv", index=False)
    print(f"  benchmark_rmssp.csv      {len(df):>6} rows")


# --------------------------------------------------------------------------
# 6. Training cost per episode
# --------------------------------------------------------------------------
def build_runtime() -> None:
    rows = []
    for src, label in METHODS.items():
        for inst in INSTANCE_ORDER:
            p = CODE / "result" / src / f"{src}_wt0.5_wc0.5_{inst}" / "log.csv"
            if not p.exists():
                continue
            d = pd.read_csv(p)
            meta = parse_instance(inst)
            rows.append({"method": label, **meta,
                         "sec_per_episode": float(d["runtime_sec"].mean()),
                         "steps_mean": float(d["steps_mean"].mean())})
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "runtime.csv", index=False)
    print(f"  runtime.csv              {len(df):>6} rows")


# --------------------------------------------------------------------------
# 7. Instance specifications
# --------------------------------------------------------------------------
def build_instances() -> None:
    rows = []
    for name in INSTANCE_ORDER:
        d = CODE / "data" / name
        meta = parse_instance(name)
        meta["cid"] = CID[name]
        based = pd.read_csv(d / "based_data.csv")
        order = pd.read_csv(d / "order_data.csv")
        proc = pd.read_csv(d / "process_data.csv")
        mod = pd.read_csv(d / "module_data.csv")
        counts = ast.literal_eval(str(order["kind_number"].iloc[0]))
        meta["n_jobs"] = int(np.sum(counts))
        meta["n_operations"] = int(np.sum(counts) * meta["J"])
        meta["n_orders"] = int(based["order_count"].iloc[0])
        # Mean routing flexibility: eligible machines / modules per operation.
        meta["avg_elig_machines"] = float(np.mean(
            [len(ast.literal_eval(str(x))) for x in proc["machine_selectable"]]))
        meta["avg_elig_modules"] = float(np.mean(
            [len(ast.literal_eval(str(x))) for x in proc["module_selectable"]]))
        # Triangular fuzzy reconfiguration times: defuzzified value and spread.
        defuzz, spread = [], []
        for col in ("time_add_fuzzy", "time_rem_fuzzy"):
            for x in mod[col]:
                l, m, u = ast.literal_eval(str(x))
                defuzz.append((l + 4 * m + u) / 6.0)
                spread.append(u - l)
        meta["avg_reconfig_time"] = float(np.mean(defuzz))
        meta["avg_fuzzy_spread"] = float(np.mean(spread))
        meta["avg_reconfig_cost"] = float(np.mean(
            list(mod["cost_add"]) + list(mod["cost_rem"])))
        rows.append(meta)
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "instances.csv", index=False)
    print(f"  instances.csv            {len(df):>6} rows")


def main() -> int:
    if not CODE.exists():
        print(f"error: code repository not found at {CODE}", file=sys.stderr)
        print("set $HCMAGRL_CODE to override", file=sys.stderr)
        return 1
    OUT.mkdir(parents=True, exist_ok=True)
    print(f"source: {CODE}\ntarget: {OUT}\n")
    build_instances()
    build_convergence()
    scored = build_eval()
    build_pareto()
    build_stats(scored)
    build_indicators()
    build_benchmark()
    build_runtime()
    print("\ndone.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
